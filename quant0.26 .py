# -*- 导出指数相关数据 -*-
import pandas as pd
import datetime
from datetime import datetime
import asyncio
import aiohttp
import requests
import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill

#先串行3000只指数股票，包括中证50，沪深300，中证500，中证1000,中证2000
#后并发2100只股票，包括中证2000和其他非指数股票
#先创建两个dataframe，第一个包含所有1900只串行股票，第二个包含3100只并发股票




wb = pd.ExcelWriter('./东财看涨看跌数据_指数分类.xlsx', engine='openpyxl',mode='r+', if_sheet_exists='replace')

Df = pd.read_excel('./东财看涨看跌数据_指数分类.xlsx',sheet_name='all_index_data')
sz50 = pd.read_excel('./东财看涨看跌数据_指数分类.xlsx',sheet_name='上证50_index_data')
hs300 = pd.read_excel('./东财看涨看跌数据_指数分类.xlsx',sheet_name='沪深300_index_data')
zz500 = pd.read_excel('./东财看涨看跌数据_指数分类.xlsx',sheet_name='中证500_index_data')
zz1000 = pd.read_excel('./东财看涨看跌数据_指数分类.xlsx',sheet_name='中证1000_index_data')
zz2000 = pd.read_excel('./东财看涨看跌数据_指数分类.xlsx',sheet_name='中证2000_index_data')


indexFile=pd.read_excel('IndexStocks.xlsx')
#indexFile包含了四列指数的数据
allIndexFile=pd.read_excel('allIndexStocks.xlsx')
#allindexFile是一列数据，包含四个指数的数据
data=pd.read_excel('stocks.xlsx',sheet_name=0)
#data是所有A股股票的数据
tickerList=data[['symbol','display_name']]
allIndex=pd.merge(allIndexFile,tickerList,on='symbol')
#该合并操作是为了在allindex的数据中加入股票名称，方便后续函数运行
allIndex=allIndex.drop_duplicates()
#去重
allIndex = allIndex.reset_index(drop=True)
#重置索引，支持后续以索引为媒介的遍历’for‘结构
allIndex=allIndex[allIndex.index<3000]
#修改参数来改变串行股票数量

indexList=[]
for i in allIndex['symbol']:
    iNumber=(i==tickerList['symbol']).argmax()
    indexList.append(iNumber)
tickerListLeft=tickerList.drop(indexList)
tickerListLeft = tickerListLeft.reset_index(drop=True)
#获取串行股票的补集

df=pd.DataFrame()
dfhs300=pd.DataFrame()
dfsz50=pd.DataFrame()
dfzz500=pd.DataFrame()
dfzz1000=pd.DataFrame()
dfzz2000=pd.DataFrame()
#生成所有数据结果，全局变量方便函数内修改DataFrame数据
count=0
async def getFromDongFangCaiFu(tickername,tickerNumber,tickerLocation):
    dataf=pd.DataFrame()
    global df,dfhs300,dfsz50,dfzz500,dfzz1000,dfzz2000, count
    async with aiohttp.ClientSession() as session:
        async with session.get('https://eminterservice.eastmoney.com/UserData/GetWebTape?code=' + tickerLocation + tickerNumber) as resp:
            print(resp.status)
            text = await resp.text()
            # print(res[:100])
            count=count+1
            print(count)
            index = text.find('TapeZ')
            if index == -1:
                dataf.loc[0,tickername] = None
            else:
                percent = text[41:48]
                index1 = percent.find(',')
                dataf.loc[0,tickername] = float(percent[:index1])
            df=pd.concat([df, dataf], axis=1)
            #查找股票所归属的指数
            if ((tickerNumber + '.' + tickerLocation) == indexFile['hs300']).sum().sum() == 1:
                dfhs300 = pd.concat([dfhs300, dataf], axis=1)
            if ((tickerNumber + '.' + tickerLocation) == indexFile['sz50']).sum().sum() == 1:
                dfsz50 = pd.concat([dfsz50, dataf], axis=1)
            if ((tickerNumber + '.' + tickerLocation) == indexFile['zz500']).sum().sum() == 1:
                dfzz500 = pd.concat([dfzz500, dataf], axis=1)
            if ((tickerNumber + '.' + tickerLocation) == indexFile['zz1000']).sum().sum() == 1:
                dfzz1000 = pd.concat([dfzz1000, dataf], axis=1)
            if ((tickerNumber + '.' + tickerLocation) == indexFile['zz2000']).sum().sum() == 1:
                dfzz2000 = pd.concat([dfzz2000, dataf], axis=1)

    return df


def CgetFromDongFangCaiFu(tickername,tickerNumber,tickerLocation):
    dataf=pd.DataFrame()
    global df,dfhs300,dfsz50,dfzz500,dfzz1000,dfzz2000
    resp=requests.get('https://eminterservice.eastmoney.com/UserData/GetWebTape?code=' + tickerLocation + tickerNumber)
    print(resp.status_code)
    text = resp.text
    # print(res[:100])
    index = text.find('TapeZ')
    if index == -1:
        dataf.loc[0,tickername] = None
    else:
        percent = text[41:48]
        index1 = percent.find(',')
        dataf.loc[0,tickername] = float(percent[:index1])
    df=pd.concat([df, dataf], axis=1)
    #查找股票所归属的指数
    if ((tickerNumber + '.' + tickerLocation) == indexFile['hs300']).sum().sum() == 1:
         dfhs300 = pd.concat([dfhs300, dataf], axis=1)
    if ((tickerNumber + '.' + tickerLocation) == indexFile['sz50']).sum().sum() == 1:
         dfsz50 = pd.concat([dfsz50, dataf], axis=1)
    if ((tickerNumber + '.' + tickerLocation) == indexFile['zz500']).sum().sum() == 1:
         dfzz500 = pd.concat([dfzz500, dataf], axis=1)
    if ((tickerNumber + '.' + tickerLocation) == indexFile['zz1000']).sum().sum() == 1:
         dfzz1000 = pd.concat([dfzz1000, dataf], axis=1)
    if ((tickerNumber + '.' + tickerLocation) == indexFile['zz2000']).sum().sum() == 1:
         dfzz2000 = pd.concat([dfzz2000, dataf], axis=1)
    return df
#串行函数

def ExcelFormatChange(sheet):
    for n in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(n)].width = 20
    for m in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for n in m:
            if type(n.value)==type(0.6):
            #检查单元格内容是否为浮点数
                if n.value > 0.6:
                    n.fill = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
                    #将大于0.6的单元格背景填充为红色
                n.number_format = '0.00%'
                #修改单元格格式为百分数，保存两位小数
            # 修改单元格格式
            # 百分数：‘0.00%’
            # 保留整数：‘0_);[Red](0)’
    return


async def main():
    global df,Df,zz50,zz500,zz1000,zz2000,hs300
    Time=(datetime.now()).strftime('%Y-%m-%d %H:%M:%S')
    df.loc[0,'日期时间']=Time
    dfsz50.loc[0, '日期时间'] = Time
    dfzz500.loc[0, '日期时间'] = Time
    dfzz1000.loc[0, '日期时间'] = Time
    dfzz2000.loc[0, '日期时间'] = Time
    dfhs300.loc[0, '日期时间'] = Time
    #pd.set_option('display.max_colwidth',10)
    taskList=[]
    global count
    t50=0
    t300=0
    t500=0
    t1000=0
    t2000=0
    startTimeForChuanXing=time.time()
    for i in range(len(allIndex)):
        t=allIndex.loc[i, 'symbol']
        name=allIndex.loc[i, 'display_name']
        tickerNumber=t[:6]
        tickerLocation=t[-2:]
        ticker=tickerNumber+tickerLocation
        tickername=name+ticker
        print(tickername)
        # x=requests.get('https://eminterservice.eastmoney.com/UserData/GetWebTape?code='+tickerLocation+tickerNumber)
        # text=x.text
        CgetFromDongFangCaiFu(tickername,tickerNumber,tickerLocation)
        count=count+1
        print(count)
        if dfsz50.shape[1]==51 and t50==0:
            zz50=zz50.append(dfsz50)
            zz50.to_excel(wb, sheet_name='上证50_index_data', index=False,)
            ExcelFormatChange(wb.sheets['上证50_index_data'])
            wb.save()
            t50=t50+1
        if dfhs300.shape[1]==301 and t300==0:
            hs300=hs300.append(dfhs300)
            hs300.to_excel(wb, sheet_name='沪深300_index_data', index=False)
            ExcelFormatChange(wb.sheets['沪深300_index_data'])
            wb.save()
            t300=t300+1
        if dfzz500.shape[1]==501 and t500==0:
            zz500=zz500.append(dfzz500)
            zz500.to_excel(wb, sheet_name='中证500_index_data', index=False)
            ExcelFormatChange(wb.sheets['中证500_index_data'])
            wb.save()
            t500=t500+1
        if dfzz1000.shape[1]==1001 and t1000==0:
            zz1000=zz1000.append(dfzz1000)
            zz1000.to_excel(wb, sheet_name='中证1000_index_data', index=False)
            ExcelFormatChange(wb.sheets['中证1000_index_data'])
            wb.save()
            t1000=t1000+1
    print("串行时间： ")
    print(time.time()-startTimeForChuanXing)
    startTimeForBingXing=time.time()
    for j in range(len(tickerListLeft)):
        t=tickerListLeft.loc[j, 'symbol']
        name=tickerListLeft.loc[j, 'display_name']
        tickerNumber=t[:6]
        tickerLocation=t[-2:]
        ticker=tickerNumber+tickerLocation
        tickername=name+ticker
        print(tickername)
        # x=requests.get('https://eminterservice.eastmoney.com/UserData/GetWebTape?code='+tickerLocation+tickerNumber)
        # text=x.text
        task=asyncio.create_task(getFromDongFangCaiFu(tickername,tickerNumber,tickerLocation))
        taskList.append(task)
    done, pending = await asyncio.wait(taskList, timeout=None)
    print("并行时间： ")
    print(time.time() - startTimeForBingXing)
    Df=Df.append(df)
    zz2000=zz2000.append(dfzz2000)
    Df.to_excel(wb,sheet_name='all_index_data',index=False)
    ExcelFormatChange(wb.sheets['all_index_data'])
    if t1000==0:
        zz1000.to_excel(wb, sheet_name='中证1000_index_data', index=False)
        ExcelFormatChange(wb.sheets['中证1000_index_data'])
    #这个判定条件是为了使得程序能够在串行股票数量小于1000的情况下正常运行
    if t2000==0:
        zz2000.to_excel(wb, sheet_name='中证2000_index_data', index=False)
        ExcelFormatChange(wb.sheets['中证2000_index_data'])
    wb.save()
    return df
if __name__ == '__main__':
    startTime=time.time()
    event_loop = asyncio.get_event_loop()
    result = event_loop.run_until_complete(asyncio.gather(main()))
    event_loop.close()
    print(result)
    print(time.time()-startTime)